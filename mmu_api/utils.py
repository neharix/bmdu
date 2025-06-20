from string import ascii_uppercase
from typing import List

import pandas as pd
from django.contrib.auth.models import User
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from rest_framework.request import HttpRequest

from main.models import Profile

from .models import Achievement, ActionLog, EducationCenter, Specialization


def is_admin(request: HttpRequest):
    return Profile.objects.get(user=request.user).role in ["superuser"]


def action_logger(request: HttpRequest, message: str):
    ActionLog.objects.create(
        user=Profile.objects.get(user=request.user), action=message
    )


def xlsx_exporter(model: str, identificators: List[int]):
    match model:
        case "profile":
            fields = ["id", "username", "password", "role", "email"]
            objects = [
                {
                    "id": item.id,
                    "username": item.user.username,
                    "password": item.password,
                    "email": item.user.email,
                    "role": item.role,
                }
                for item in Profile.objects.filter(id__in=identificators)
            ]
        case "education-center":
            fields = [
                "id",
                "name",
                "is_active",
                "phone_number",
                "lat",
                "lng",
                "country",
                "region",
                "address",
                "buildings_count",
                "rooms_count",
                "capacity",
                "books_count",
            ]
            objects = [
                {
                    "id": item.id,
                    "name": item.name,
                    "is_active": item.is_active,
                    "phone_number": item.phone_number,
                    "lat": item.lat,
                    "lng": item.lng,
                    "country": item.country.id,
                    "region": item.region.id,
                    "address": item.address,
                    "buildings_count": item.buildings_count,
                    "rooms_count": item.rooms_count,
                    "capacity": item.capacity,
                    "books_count": item.books_count,
                }
                for item in EducationCenter.objects.filter(id__in=identificators)
            ]
        case "achievement":
            fields = ["id", "name"]
            objects = [
                {"id": item.id, "name": item.name}
                for item in Achievement.objects.filter(id__in=identificators)
            ]
        case "specialization":
            fields = ["id", "name"]
            objects = [
                {"id": item.id, "name": item.name}
                for item in Specialization.objects.filter(id__in=identificators)
            ]

    row_count = len(objects)

    wb = Workbook()
    wb.remove(wb.worksheets[0])

    font = Font(name="Times New Roman", size=11, color="000000")
    alignment = Alignment(
        horizontal="center", vertical="center", wrap_text=True, wrapText=True
    )
    bold_font = Font(name="Times New Roman", size=11, color="000000", bold=True)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    ws: Worksheet = wb.create_sheet(model.capitalize())

    column_index = 0

    for column_id in ascii_uppercase[: len(fields)]:
        ws[f"{column_id}1"].value = fields[column_index].replace("_", " ").capitalize()
        column_index += 1
        ws.column_dimensions[column_id].width = 20
        for row_id in range(1, row_count + 2):
            ws[f"{column_id}{row_id}"].border = border
            ws[f"{column_id}{row_id}"].alignment = alignment
            ws[f"{column_id}{row_id}"].font = bold_font if row_id == 1 else font
            ws.row_dimensions[row_id].height = 50

    if row_count > 0:
        for row_id in range(2, row_count + 2):
            for field in fields:
                ws[f"{ascii_uppercase[fields.index(field)]}{row_id}"].value = objects[
                    row_id - 2
                ].get(field)

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 20

    return wb


def xlsx_importer(model: str, excel):  # excel : request.FILES["excel"]
    match model:
        case "profile":
            dataframe = pd.read_excel(excel)
            for index, row in dataframe.iterrows():
                username = row["Username"].lstrip().rstrip()
                password = row["Password"].lstrip().rstrip()
                email = row["Email"].lstrip().rstrip()
                role = row["Role"].lstrip().rstrip()
                user = User.objects.create_user(
                    username=username, password=password, email=email
                )
                profile = Profile.objects.get(user=user)
                profile.password = password
                profile.allowed_service = "mmu"
                profile.role = role
                profile.save()
        case _:
            print("Doesn't found model")
